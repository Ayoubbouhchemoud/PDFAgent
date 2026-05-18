"""
pdf_to_xlsx.py — PDF to Excel (.xlsx) converter.

Reuses the same pdfplumber-based structural extraction as pdf_to_docx.py
and maps each logical block to openpyxl cells with a 10-column grid:

  Full-width content  → merged A:J
  Left code box (40%) → merged A:D
  Right code box (60%)→ merged E:J  (matches ~40/60% PDF column ratio)
  Bullet items        → A = "•", B:J merged = text
  Header logo         → floating image at A (cols A:B area), text at C:J
  Page separator      → medium-border row between pages
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

_HERE = Path(__file__).parent
sys.path.insert(0, str(_HERE))

from pdf_to_html import (
    _COL_DETECT_GAP,
    _LINE_TOL,
    _MERGE_GAP,
    _MIN_TABLE_COLS,
    _assign_col,
    _base_font,
    _content_boxes,
    _detect_col_starts,
    _group_into_lines,
    _is_bold,
    _is_italic,
    _is_mono,
    _is_white_color,
    _merge_into_runs,
    _remap_pua,
    _resolve_indexed_palette,
    _section_rects,
)
from pdf_to_docx import (
    _extract_image_bytes,
    _find_box_pairs,
    _group_into_paragraphs,
    _is_bullet_word,
    _line_height,
    _map_fontname,
    _underline_ys,
)

# ---------------------------------------------------------------------------
# Grid constants
# ---------------------------------------------------------------------------
N_COLS = 10           # columns A–J
COL_W  = 12.0         # character width per column

# Code-box column split: left=A:D (40%), right=E:J (60%)
# Matches the PDF's ~40/60% proportions across all three box pairs.
COL_LEFT_END    = 4   # left box end column (inclusive), A:D
COL_RIGHT_START = 5   # right box start column, E:J

# Header area: image floats at A:B, text starts at C
COL_HDR_IMG_END   = 2
COL_HDR_TEXT_START = 3

COL_BULLET_CHAR = 1   # col A: bullet glyph
COL_BULLET_TEXT = 2   # cols B–J: bullet text

ROW_H_NORMAL  = 15.0
ROW_H_HEADING = 15.0
ROW_H_CODE    = 13.0
ROW_H_SPACER  =  5.0
ROW_H_LABEL   = 14.0

_THIN = Side(border_style="thin",   color="000000")
_MED  = Side(border_style="medium", color="AAAAAA")
_NO   = Side(border_style=None)


# ---------------------------------------------------------------------------
# Text helpers
# ---------------------------------------------------------------------------

def _words_to_text(words: list[dict]) -> str:
    if not words:
        return ""
    parts: list[str] = []
    prev_x1: float | None = None
    for w in sorted(words, key=lambda x: x["x0"]):
        text = _remap_pua(w["text"], w.get("fontname", ""))
        if prev_x1 is not None and w["x0"] - prev_x1 > 1.0:
            parts.append(" ")
        parts.append(text)
        prev_x1 = w.get("x1", w["x0"])
    return "".join(parts)


def _line_to_text(line: list[dict]) -> str:
    return _words_to_text(line)


def _para_to_text(para_lines: list[list[dict]]) -> str:
    return " ".join(_line_to_text(ln) for ln in para_lines if ln)


def _detect_font(words: list[dict]) -> tuple[str, float, bool, bool]:
    """Return (excel_font_name, size_pt, bold, italic) from the first word."""
    if not words:
        return "Calibri", 11.0, False, False
    w = words[0]
    fname = w.get("fontname", "")
    return (
        _map_fontname(fname),
        float(w.get("size", 11)),
        _is_bold(fname),
        _is_italic(fname),
    )


# ---------------------------------------------------------------------------
# Sheet writer
# ---------------------------------------------------------------------------

class SheetWriter:
    """Row-by-row writer for a single openpyxl worksheet."""

    def __init__(self, ws):
        self.ws  = ws
        self.row = 1
        for c in range(1, N_COLS + 1):
            ws.column_dimensions[get_column_letter(c)].width = COL_W

    # ── Primitives ────────────────────────────────────────────────────────

    def _cell(self, col: int):
        return self.ws.cell(row=self.row, column=col)

    def _merge(self, col_start: int, col_end: int) -> None:
        if col_end > col_start:
            self.ws.merge_cells(
                start_row=self.row, start_column=col_start,
                end_row=self.row,   end_column=col_end,
            )

    def _set_height(self, h: float) -> None:
        self.ws.row_dimensions[self.row].height = h

    # ── Public writers ───────────────────────────────────────────────────

    def spacer(self, height: float = ROW_H_SPACER) -> None:
        self._set_height(height)
        self.row += 1

    def text_row(
        self,
        text: str,
        bold: bool = False,
        italic: bool = False,
        font_name: str = "Calibri",
        size: float = 11.0,
        col_start: int = 1,
        col_end: int = N_COLS,
        bottom_border: bool = False,
        top_border: bool = False,
        height: float = ROW_H_NORMAL,
        h_align: str = "left",
        wrap: bool = True,
    ) -> None:
        self._merge(col_start, col_end)
        cell = self._cell(col_start)
        cell.value = text
        cell.font  = Font(name=font_name, size=size, bold=bold, italic=italic)
        cell.alignment = Alignment(
            horizontal=h_align, vertical="center", wrap_text=wrap)
        b_top    = _THIN if top_border    else _NO
        b_bottom = _THIN if bottom_border else _NO
        if top_border or bottom_border:
            cell.border = Border(top=b_top, bottom=b_bottom)
        self._set_height(height)
        self.row += 1

    def bullet_row(
        self,
        text: str,
        font_name: str = "Calibri",
        size: float = 11.0,
        height: float = ROW_H_NORMAL,
    ) -> None:
        bullet_cell = self._cell(COL_BULLET_CHAR)
        bullet_cell.value = "•"
        bullet_cell.font  = Font(name=font_name, size=size)
        bullet_cell.alignment = Alignment(
            horizontal="center", vertical="center")

        self._merge(COL_BULLET_TEXT, N_COLS)
        text_cell = self._cell(COL_BULLET_TEXT)
        text_cell.value = text
        text_cell.font  = Font(name=font_name, size=size)
        text_cell.alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True)

        self._set_height(height)
        self.row += 1

    def _code_cell(
        self,
        text: str,
        col_start: int,
        col_end: int,
        is_first: bool,
        is_last: bool,
        font_name: str = "Consolas",
        size: float = 10.0,
    ) -> None:
        self._merge(col_start, col_end)
        cell = self._cell(col_start)
        cell.value = text
        cell.font  = Font(name=font_name, size=size)
        cell.alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=False)
        cell.border = Border(
            left=_THIN,
            right=_THIN,
            top=(_THIN if is_first else _NO),
            bottom=(_THIN if is_last  else _NO),
        )

    def code_rows(
        self,
        lines: list[str],
        col_start: int = 1,
        col_end: int = N_COLS,
        font_name: str = "Consolas",
        size: float = 10.0,
    ) -> None:
        n = len(lines)
        if n == 0:
            return
        for i, text in enumerate(lines):
            self._code_cell(text, col_start, col_end,
                            is_first=(i == 0), is_last=(i == n - 1),
                            font_name=font_name, size=size)
            self._set_height(ROW_H_CODE)
            self.row += 1

    def side_by_side_rows(
        self,
        left_lines:  list[str],
        right_lines: list[str],
        left_font:   str = "Consolas",
        right_font:  str = "Consolas",
        size: float = 10.0,
    ) -> None:
        """Write two code boxes side-by-side: left=A:D, right=E:J."""
        n = max(len(left_lines), len(right_lines), 1)
        left_padded  = left_lines  + [""] * (n - len(left_lines))
        right_padded = right_lines + [""] * (n - len(right_lines))
        for i, (lt, rt) in enumerate(zip(left_padded, right_padded)):
            is_first = (i == 0)
            is_last  = (i == n - 1)
            self._code_cell(lt, 1, COL_LEFT_END,    is_first, is_last,
                            font_name=left_font,  size=size)
            self._code_cell(rt, COL_RIGHT_START, N_COLS, is_first, is_last,
                            font_name=right_font, size=size)
            self._set_height(ROW_H_CODE)
            self.row += 1

    def label_row(
        self,
        left_text:  str,
        right_text: str,
        font_name:  str = "Calibri",
        size: float = 11.0,
        bold: bool = False,
        height: float = ROW_H_LABEL,
    ) -> None:
        if not left_text and not right_text:
            return
        # Left label: spans the full left-box width (A:D)
        self._merge(1, COL_LEFT_END)
        lc = self._cell(1)
        lc.value = left_text
        lc.font  = Font(name=font_name, size=size, bold=bold)
        lc.alignment = Alignment(horizontal="left", vertical="center")

        # Right label: spans the full right-box width (E:J)
        self._merge(COL_RIGHT_START, N_COLS)
        rc = self._cell(COL_RIGHT_START)
        rc.value = right_text
        rc.font  = Font(name=font_name, size=size, bold=bold)
        rc.alignment = Alignment(horizontal="left", vertical="center")

        self._set_height(height)
        self.row += 1

    def place_image(
        self,
        raw_bytes: bytes,
        img_w_pt: float,
        img_h_pt: float,
        anchor_row: int,
    ) -> None:
        """Embed a PNG image anchored at col A of anchor_row (floating)."""
        try:
            from openpyxl.drawing.image import Image as XLImage
            xl_img = XLImage(io.BytesIO(raw_bytes))
            xl_img.width  = max(int(img_w_pt * 96 / 72), 10)
            xl_img.height = max(int(img_h_pt * 96 / 72), 10)
            xl_img.anchor = f"A{anchor_row}"
            self.ws.add_image(xl_img)
        except Exception:
            pass

    def standalone_image_row(
        self,
        raw_bytes: bytes,
        img_w_pt: float,
        img_h_pt: float,
    ) -> None:
        """Write a stand-alone image in its own row (non-header images)."""
        row_h = max(img_h_pt * 0.75, 20.0)
        self.place_image(raw_bytes, img_w_pt, img_h_pt, self.row)
        self._set_height(row_h)
        self.row += 1

    def page_separator(self) -> None:
        self.spacer(height=6.0)
        self._merge(1, N_COLS)
        sep = self._cell(1)
        sep.border = Border(bottom=_MED)
        self._set_height(3.0)
        self.row += 1
        self.spacer(height=6.0)

    def table_rows(
        self,
        words: list[dict],
        col_starts: list[float],
        font_name: str = "Calibri",
        size: float = 10.0,
    ) -> None:
        lines = _group_into_lines(words)
        if not lines:
            return
        n_data = len(col_starts)
        xl_per_data = max(1, N_COLS // n_data)

        for line in lines:
            all_bold_line = all(_is_bold(w.get("fontname", "")) for w in line)
            for ci in range(n_data):
                c_words = [w for w in line
                           if _assign_col(w["x0"], col_starts) == ci]
                text = _words_to_text(c_words)
                xl_cs = 1 + ci * xl_per_data
                xl_ce = xl_cs + xl_per_data - 1
                if ci == n_data - 1:
                    xl_ce = N_COLS
                self._merge(xl_cs, xl_ce)
                cell = self._cell(xl_cs)
                cell.value = text
                cell.font  = Font(name=font_name, size=size,
                                  bold=all_bold_line)
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True)
                cell.border = Border(
                    left=_THIN, right=_THIN,
                    top=_THIN, bottom=_THIN)
            self._set_height(ROW_H_NORMAL)
            self.row += 1


# ---------------------------------------------------------------------------
# Per-page processing
# ---------------------------------------------------------------------------

def _process_page(sw: SheetWriter, page, page_idx: int) -> None:  # noqa: ARG001
    pw = float(page.width)
    ph = float(page.height)

    # Detect footer separator: widest horizontal rect in the bottom 20% of the page.
    footer_y = ph
    for rect in (page.rects or []):
        rh = float(rect["bottom"]) - float(rect["top"])
        rw = float(rect["x1"])    - float(rect["x0"])
        if rh < 2 and rw > pw * 0.5 and float(rect["top"]) > ph * 0.8:
            footer_y = min(footer_y, float(rect["top"]))

    all_words = [
        w for w in page.extract_words(extra_attrs=["size", "fontname"])
        if float(w["top"]) < footer_y
    ]

    # ── Content boxes ─────────────────────────────────────────────────────
    boxes = _content_boxes(page)

    box_words: dict[int, list[dict]] = {i: [] for i in range(len(boxes))}
    free_words: list[dict] = []
    for w in all_words:
        wx1 = w.get("x1", w["x0"] + 1)
        wb  = w.get("bottom", w["top"] + 1)
        assigned = False
        for i, box in enumerate(boxes):
            if (box["x0"] - 2 <= w["x0"] and wx1 <= box["x1"] + 2 and
                    box["top"] - 2 <= w["top"] and wb <= box["bottom"] + 2):
                box_words[i].append(w)
                assigned = True
                break
        if not assigned:
            free_words.append(w)

    # ── Table detection ───────────────────────────────────────────────────
    sections = _section_rects(page, bordered_boxes=boxes)
    table_regions: list[tuple] = []
    for (y_top, y_bot) in sections:
        band = [w for w in free_words if y_top - 1 <= w["top"] <= y_bot + 1]
        if not band:
            continue
        lines = _group_into_lines(band)
        if not lines:
            continue
        cs = _detect_col_starts(lines[0])
        if len(cs) >= _MIN_TABLE_COLS:
            table_regions.append((y_top, y_bot, cs))

    non_table_free = [
        w for w in free_words
        if not any(yt - 1 <= w["top"] <= yb + 1 for (yt, yb, _cs) in table_regions)
    ]

    # ── Heading underlines ────────────────────────────────────────────────
    ul_ys = _underline_ys(page)

    def _near_underline(top: float, bottom: float) -> bool:
        return any(top - 2 <= y <= bottom + 20 for y in ul_ys)

    # ── Box pairing ───────────────────────────────────────────────────────
    pairs, singles = _find_box_pairs(boxes)

    def _labels_for_pair(lbox: dict, rbox: dict) -> tuple[str, str, str, float, bool]:
        """Split label words at the actual gap between the two boxes.

        Returns (left_text, right_text, font_name, size, bold).
        """
        box_top = min(float(lbox["top"]), float(rbox["top"]))
        gap_mid = (float(lbox["x1"]) + float(rbox["x0"])) / 2
        candidates = [
            w for w in free_words
            if 0 < box_top - w["top"] < 35 and
               "symbol" not in w.get("fontname", "").lower()
        ]
        if not candidates:
            return "", "", "Calibri", 11.0, False
        ls = _group_into_lines(candidates)
        if not ls:
            return "", "", "Calibri", 11.0, False
        last_line = ls[-1]
        fn, sz, bold, _ = _detect_font(last_line)
        ll = " ".join(
            _remap_pua(w["text"], w.get("fontname", ""))
            for w in last_line if w["x0"] < gap_mid)
        rl = " ".join(
            _remap_pua(w["text"], w.get("fontname", ""))
            for w in last_line if w["x0"] >= gap_mid)
        return ll, rl, fn, sz, bold

    def _label_above_single(box_top: float) -> str:
        candidates = [
            w for w in free_words
            if 0 < box_top - w["top"] < 35 and
               "symbol" not in w.get("fontname", "").lower()
        ]
        if not candidates:
            return ""
        ls = _group_into_lines(candidates)
        if not ls:
            return ""
        return " ".join(
            _remap_pua(w["text"], w.get("fontname", ""))
            for w in ls[-1])

    # ── Label-line exclusion ──────────────────────────────────────────────
    label_line_ys: set[float] = set()
    for (lbox, rbox) in pairs:
        for w in free_words:
            if 0 < lbox["top"] - w["top"] < 35:
                label_line_ys.add(round(w["top"], 1))
    for box in singles:
        for w in free_words:
            if 0 < box["top"] - w["top"] < 35:
                label_line_ys.add(round(w["top"], 1))

    # ── Paragraph groups ──────────────────────────────────────────────────
    content_words = [
        w for w in non_table_free
        if round(w["top"], 1) not in label_line_ys
    ]
    content_lines = _group_into_lines(content_words)
    para_groups   = _group_into_paragraphs(content_lines, pw)

    # ── Image → header-layout or standalone ──────────────────────────────
    images = list(page.images)
    consumed_para_ids: set[int] = set()

    image_events: list[tuple[float, str, object]] = []
    for img in images:
        img_top = float(img["top"])
        img_bot = float(img["bottom"])
        img_x1  = float(img["x1"])
        side_paras = [
            pg for pg in para_groups
            if (pg[0][0]["top"] < img_bot + 5 and
                max(w.get("bottom", w["top"] + 12) for ln in pg for w in ln) > img_top - 5 and
                min(w["x0"] for ln in pg for w in ln) > img_x1 - 5)
        ]
        if side_paras:
            for pg in side_paras:
                consumed_para_ids.add(id(pg))
            image_events.append((img_top, "header", (img, side_paras)))
        else:
            image_events.append((img_top, "image", img))

    # ── Build event list ──────────────────────────────────────────────────
    rendered_table: set = set()
    rendered_box:   set = set()
    rendered_pair:  set = set()

    events: list[tuple[float, str, object]] = list(image_events)
    for (yt, yb, cs) in table_regions:
        events.append((yt, "table", (yt, yb, cs)))
    for pair_idx, (lbox, rbox) in enumerate(pairs):
        events.append((min(lbox["top"], rbox["top"]), "pair", pair_idx))
    for box_idx, box in enumerate(singles):
        events.append((box["top"], "single", box_idx))
    for pg in para_groups:
        events.append((pg[0][0]["top"], "para", pg))
    events.sort(key=lambda e: e[0])

    # ── Emit rows in reading order ────────────────────────────────────────
    bullet_buffer: list[list[dict]] = []

    def flush_bullets() -> None:
        for line in bullet_buffer:
            text_words = [w for w in line if not _is_bullet_word(w)]
            fn, sz, _, _ = _detect_font(text_words)
            sw.bullet_row(_words_to_text(text_words), font_name=fn, size=sz)
        bullet_buffer.clear()

    for (y, kind, data) in events:

        if kind == "header":
            flush_bullets()
            img, side_paras = data
            result = _extract_image_bytes(img)
            img_w_pt = float(img["x1"]) - float(img["x0"])
            img_h_pt = float(img["bottom"]) - float(img["top"])
            n_rows   = max(len(side_paras), 1)
            row_h    = max(img_h_pt / n_rows, ROW_H_NORMAL)
            anchor_row = sw.row

            # Write header text at C:J — leaves A:B for the floating image
            for i, para_lines in enumerate(side_paras):
                text = _para_to_text(para_lines)
                all_words_pg = [w for ln in para_lines for w in ln]
                fn, sz, bold, italic = _detect_font(all_words_pg)
                is_last = (i == n_rows - 1)
                sw.text_row(
                    text, bold=bold, italic=italic,
                    font_name=fn, size=sz,
                    col_start=COL_HDR_TEXT_START, col_end=N_COLS,
                    height=row_h,
                    bottom_border=is_last,   # underline after last header row
                )

            # Place image floating at A{anchor_row}, beside the text
            if result:
                raw, _ = result
                sw.place_image(raw, img_w_pt, img_h_pt, anchor_row)

        elif kind == "image":
            flush_bullets()
            result = _extract_image_bytes(data)
            if result:
                raw, _ = result
                img_w_pt = float(data["x1"]) - float(data["x0"])
                img_h_pt = float(data["bottom"]) - float(data["top"])
                sw.standalone_image_row(raw, img_w_pt, img_h_pt)

        elif kind == "table":
            (yt, yb, cs) = data
            if (yt, yb) in rendered_table:
                continue
            rendered_table.add((yt, yb))
            flush_bullets()
            tbl_words = [w for w in free_words if yt - 1 <= w["top"] <= yb + 1]
            if tbl_words:
                sw.table_rows(tbl_words, cs)
            sw.spacer()

        elif kind == "pair":
            pair_idx = data
            if pair_idx in rendered_pair:
                continue
            rendered_pair.add(pair_idx)
            flush_bullets()
            lbox, rbox = pairs[pair_idx]
            ll, rl, lbl_fn, lbl_sz, lbl_bold = _labels_for_pair(lbox, rbox)

            l_words = box_words[boxes.index(lbox)]
            r_words = box_words[boxes.index(rbox)]

            l_lines = [_line_to_text(ln) for ln in _group_into_lines(l_words)]
            r_lines = [_line_to_text(ln) for ln in _group_into_lines(r_words)]

            l_fn, l_sz, _, _ = _detect_font(l_words)
            r_fn, r_sz, _, _ = _detect_font(r_words)

            sw.label_row(ll, rl, font_name=lbl_fn, size=lbl_sz, bold=lbl_bold)
            sw.side_by_side_rows(l_lines, r_lines,
                                  left_font=l_fn, right_font=r_fn,
                                  size=min(l_sz, r_sz))
            sw.spacer()

        elif kind == "single":
            box_idx = data
            if box_idx in rendered_box:
                continue
            rendered_box.add(box_idx)
            flush_bullets()
            box   = singles[box_idx]
            words = box_words[boxes.index(box)]
            lines = [_line_to_text(ln) for ln in _group_into_lines(words)]
            fn, sz, _, _ = _detect_font(words)
            sw.code_rows(lines, font_name=fn, size=sz)
            sw.spacer()

        elif kind == "para":
            pg: list[list[dict]] = data
            if id(pg) in consumed_para_ids:
                continue

            first_line = pg[0]

            if first_line and _is_bullet_word(first_line[0]):
                for line in pg:
                    bullet_buffer.append(line)
                continue

            flush_bullets()

            all_bold_line = all(_is_bold(w.get("fontname", "")) for w in first_line)
            top_y = first_line[0]["top"]
            bot_y = max(w.get("bottom", w["top"] + 12) for w in first_line)
            is_underlined = _near_underline(top_y, bot_y)

            text = _para_to_text(pg)
            all_words_pg = [w for ln in pg for w in ln]
            fn, sz, bold, italic = _detect_font(all_words_pg)

            # Row height: one slot per visual line that was merged
            n_visual_lines = len(pg)
            row_h = n_visual_lines * ROW_H_NORMAL

            if all_bold_line and is_underlined:
                sw.spacer(height=4.0)
                sw.text_row(
                    text, bold=True, italic=italic,
                    font_name=fn, size=sz,
                    bottom_border=True,
                    height=ROW_H_HEADING,
                )
            else:
                sw.text_row(
                    text, bold=bold, italic=italic,
                    font_name=fn, size=sz,
                    height=row_h,
                )

    flush_bullets()


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def convert(pdf_path: str | Path, title: str | None = None) -> bytes:
    """Convert a PDF to XLSX and return the raw bytes."""
    pdf_path    = Path(pdf_path)
    sheet_title = (title or pdf_path.stem)[:31]

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    sw = SheetWriter(ws)

    with pdfplumber.open(pdf_path) as pdf:
        for page_idx, page in enumerate(pdf.pages):
            if page_idx > 0:
                sw.page_separator()
            _process_page(sw, page, page_idx)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys as _sys

    if len(_sys.argv) < 2:
        print("Usage: python pdf_to_xlsx.py <input.pdf> [output.xlsx]",
              file=_sys.stderr)
        _sys.exit(1)

    src = Path(_sys.argv[1])
    dst = Path(_sys.argv[2]) if len(_sys.argv) > 2 else src.with_suffix(".xlsx")
    dst.write_bytes(convert(src))
    print(f"Written → {dst}")
